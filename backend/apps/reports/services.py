import io
import csv
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from django.db.models import Sum, Count, Avg
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

class ReportGenerator:
    
    @staticmethod
    def generate_order_report(start_date, end_date, format='pdf'):
        orders = Order.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).select_related('customer', 'vendor', 'restaurant')
        
        if format == 'pdf':
            return ReportGenerator._generate_pdf_report(orders, start_date, end_date)
        elif format == 'excel':
            return ReportGenerator._generate_excel_report(orders, start_date, end_date)
        elif format == 'csv':
            return ReportGenerator._generate_csv_report(orders, start_date, end_date)
    
    @staticmethod
    def _generate_pdf_report(orders, start_date, end_date):
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#2E3B4E'),
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        title = Paragraph(f"Order Report: {start_date.date()} to {end_date.date()}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))
        
        # Summary section
        summary_data = [
            ['Total Orders', 'Total Revenue', 'Average Order Value', 'Completed Orders', 'Cancelled Orders'],
            [
                orders.count(),
                f"${orders.aggregate(total=Sum('total_amount'))['total'] or 0:.2f}",
                f"${orders.aggregate(avg=Avg('total_amount'))['avg'] or 0:.2f}",
                orders.filter(status='DELIVERED').count(),
                orders.filter(status='CANCELLED').count()
            ]
        ]
        
        summary_table = Table(summary_data, colWidths=[1.2 * inch] * 5)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3 * inch))
        
        # Order details table
        data = [['Order #', 'Customer', 'Vendor', 'Amount', 'Status', 'Date']]
        for order in orders[:100]:  # Limit to 100 orders for PDF
            data.append([
                order.order_number,
                order.customer.full_name,
                order.vendor.business_name,
                f"${order.total_amount}",
                order.status,
                order.created_at.strftime('%Y-%m-%d %H:%M')
            ])
        
        order_table = Table(data, colWidths=[1.5 * inch, 1.5 * inch, 1.5 * inch, 1 * inch, 1 * inch, 1.5 * inch])
        order_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
        ]))
        
        elements.append(order_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="order_report_{start_date.date()}_to_{end_date.date()}.pdf"'
        return response
    
    @staticmethod
    def _generate_excel_report(orders, start_date, end_date):
        wb = openpyxl.Workbook()
        
        # Orders sheet
        ws = wb.active
        ws.title = "Orders"
        
        # Headers
        headers = ['Order Number', 'Customer', 'Vendor', 'Restaurant', 'Subtotal', 'Tax', 'Delivery Fee', 'Total', 'Status', 'Date']
        ws.append(headers)
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Add data
        for order in orders:
            ws.append([
                order.order_number,
                order.customer.full_name,
                order.vendor.business_name,
                order.restaurant.name,
                float(order.subtotal),
                float(order.tax_amount),
                float(order.delivery_fee),
                float(order.total_amount),
                order.status,
                order.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Summary sheet
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(['Metric', 'Value'])
        summary_ws.append(['Total Orders', orders.count()])
        summary_ws.append(['Total Revenue', f"${orders.aggregate(total=Sum('total_amount'))['total'] or 0:.2f}"])
        summary_ws.append(['Average Order Value', f"${orders.aggregate(avg=Avg('total_amount'))['avg'] or 0:.2f}"])
        summary_ws.append(['Completed Orders', orders.filter(status='DELIVERED').count()])
        summary_ws.append(['Cancelled Orders', orders.filter(status='CANCELLED').count()])
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="order_report_{start_date.date()}_to_{end_date.date()}.xlsx"'
        wb.save(response)
        return response
    
    @staticmethod
    def _generate_csv_report(orders, start_date, end_date):
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="order_report_{start_date.date()}_to_{end_date.date()}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Order Number', 'Customer', 'Vendor', 'Restaurant', 'Subtotal', 'Tax', 'Delivery Fee', 'Total', 'Status', 'Date'])
        
        for order in orders:
            writer.writerow([
                order.order_number,
                order.customer.full_name,
                order.vendor.business_name,
                order.restaurant.name,
                float(order.subtotal),
                float(order.tax_amount),
                float(order.delivery_fee),
                float(order.total_amount),
                order.status,
                order.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return response